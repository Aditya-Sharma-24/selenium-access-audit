# Web Scraping Tool for User Access Audits - Aditya Sharma, Aug 2024.
def script_option_1():
    import openpyxl
    from selenium.webdriver import Edge
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    import tkinter as tk
    from tkinter import messagebox

    def wait_for_user_navigation():
        root = tk.Tk()
        root.withdraw()
        messagebox.showinfo("Manual Navigation Required", "Please log in and navigate to the required page. Click OK when done.")
        root.destroy()

    # Replace this with your actual URL where the user will log in manually
    login_url = "http://www.internal-tool-url.com/"
    data_rows_selector = "tbody tr[data-testid='p-os-generic-table-row']"
    next_page_button = "tfoot a[type='nextItem']"

    # Initialize the Edge WebDriver
    driver = Edge()
    driver.get(login_url)

    # Wait for the user to manually log in
    wait_for_user_navigation()

    # Function to scrape data from the current page and write to Excel
    def scrape_and_write_data():
        wait = WebDriverWait(driver, 10)
        rows = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, data_rows_selector)))
        for row in rows:
            data = [cell.text.strip() for cell in row.find_elements(By.TAG_NAME, "td")]
            sheet.append(data)

    # Initialize Excel workbook and sheet
    wb = openpyxl.Workbook()
    sheet = wb.active
    next_button_enabled = True

    while next_button_enabled:
        scrape_and_write_data()
        next_button = driver.find_element(By.CSS_SELECTOR, next_page_button)
        next_button_enabled = "disabled" not in next_button.get_attribute("class")
        if next_button_enabled:
            next_button.click()
            driver.implicitly_wait(4)

    wb.save("scraped_data.xlsx")
    driver.quit()
def script_option_2():
    # Full script from Code 2
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.edge.service import Service
    from selenium.webdriver.edge.options import Options
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.common.exceptions import NoSuchElementException, TimeoutException
    import openpyxl
    import time
    import os

    # Constants
    INPUT_EXCEL_PATH = "scraped_data.xlsx"
    OUTPUT_EXCEL_PATH = "User_Details.xlsx"
    URL = "http://www.internal-tool-url.com/"
    IMPLICIT_WAIT = 3
    DATA_CHECK_INTERVAL = 5

    def read_user_groups(input_path):
        """Reads user group names from the Excel file."""
        workbook = openpyxl.load_workbook(input_path)
        sheet = workbook.active
        user_groups = [sheet.cell(row=i, column=2).value for i in range(2, sheet.max_row + 1)]
        workbook.close()
        return user_groups

    def append_user_details(output_path, data):
        """Appends user details to an existing Excel file or creates one if it doesn't exist."""
        file_exists = os.path.exists(output_path)
        if file_exists:
            workbook = openpyxl.load_workbook(output_path)
            sheet = workbook.active
        else:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["First Name", "Last Name", "Email", "Phone", "Role", "User Group"])

        for row in data:
            sheet.append(row)
        workbook.save(output_path)

    def scrape_user_details(driver, user_group):
        """Scrapes user details from the current table, handles pagination."""
        data = []
        last_page_data = None
        while True:
            # Locate the user details table and scrape data
            rows = driver.find_elements(By.XPATH, "//div[@class='user-search']//tr")
            current_page_data = []
            for row in rows[1:]:
                cols = row.find_elements(By.TAG_NAME, "td")
                if cols:
                    first_name = cols[0].text
                    last_name = cols[1].text
                    email = cols[2].text
                    phone = cols[3].text
                    role = cols[4].text
                    current_page_data.append([first_name, last_name, email, phone, role, user_group])

            # Check if data has changed
            if current_page_data == last_page_data:
                time.sleep(DATA_CHECK_INTERVAL)
                if current_page_data == last_page_data:
                    break
            else:
                last_page_data = current_page_data
                data.extend(current_page_data)

            # Check if "Next" button is enabled
            try:
                next_button = driver.find_element(By.XPATH, "//i[@class='icon pos arrow-right']")
                if "disabled" in next_button.get_attribute("class"):
                    break
                next_button.click()

            except:
                break
        return data

    def open_all_user_groups(driver, user_group):
        """Opens all user groups listed in search results and scrapes details."""
        user_data = []
        results = driver.find_elements(By.XPATH, "//i[@class='icon pos view']")
        for i in range(len(results)):
            # Refresh the list of results
            results = driver.find_elements(By.XPATH, "//i[@class='icon pos view']")
            results[i].click()

            # Scrape user details, including pagination
            user_data.extend(scrape_user_details(driver, user_group))

            # Go back to the main page
            back_button = driver.find_element(By.XPATH, "//a[normalize-space()='Authorizer']")
            back_button.click()
            time.sleep(1)

        return user_data

    def main():
        while True:
            try:
                # Load user groups from Excel
                try:
                    user_groups = read_user_groups(INPUT_EXCEL_PATH)
                except FileNotFoundError:
                    print(f"Error: The file '{INPUT_EXCEL_PATH}' was not found.")
                    return

                # Initialize the Edge WebDriver
                options = Options()
                options.add_argument("start-maximized")
                service = Service("path/to/msedgedriver.exe")
                driver = webdriver.Edge(service=service, options=options)
                driver.implicitly_wait(IMPLICIT_WAIT)

                try:
                    driver.get(URL)

                    # Click on the User Groups button to start the script automatically
                    user_groups_button = WebDriverWait(driver, IMPLICIT_WAIT).until(
                        EC.presence_of_element_located((By.XPATH, "//a[normalize-space()='User Groups']"))
                    )
                    user_groups_button.click()

                    total_user_groups = len(user_groups)
                    for index, user_group in enumerate(user_groups):
                        print(f"{index + 1} out of {total_user_groups} user groups scraped")

                        # Clear the search box before entering a new user group name
                        search_box_found = False
                        while not search_box_found:
                            try:
                                search_box = driver.find_element(By.XPATH,
                                                                 "//input[@placeholder='Search User Group by Name or ID']")
                                search_box_found = True
                            except NoSuchElementException:
                                time.sleep(0)

                        search_box.click()
                        search_box.send_keys(Keys.CONTROL, "a")
                        search_box.send_keys(Keys.BACKSPACE)

                        # Enter the new user group name
                        search_box.send_keys(user_group)
                        search_box.send_keys(Keys.RETURN)
                        time.sleep(0)

                        # Open all user groups listed in search results and scrape details
                        user_data = open_all_user_groups(driver, user_group)
                        append_user_details(OUTPUT_EXCEL_PATH, user_data)

                finally:
                    driver.quit()
                    print("Scraping complete. Data saved to", OUTPUT_EXCEL_PATH)

                break  # Exit the loop if no issues occur

            except (NoSuchElementException, TimeoutException) as e:
                print(f"Error occurred: {e}. Restarting script...")

    if __name__ == "__main__":
        main()

    pass

def script_option_3():
    # Full script from Code 3
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.edge.service import Service
    from selenium.webdriver.edge.options import Options
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.common.exceptions import NoSuchElementException, TimeoutException
    import openpyxl
    import time
    import os

    # Constants
    INPUT_EXCEL_PATH = "scraped_data.xlsx"
    OUTPUT_EXCEL_PATH = "User_Details.xlsx"
    URL = "http://www.internal-tool-url.com/"
    IMPLICIT_WAIT = 8
    DATA_CHECK_INTERVAL = 15

    def read_user_groups(input_path):
        """Reads user group names from the Excel file."""
        workbook = openpyxl.load_workbook(input_path)
        sheet = workbook.active
        user_groups = [sheet.cell(row=i, column=2).value for i in range(2, sheet.max_row + 1)]
        workbook.close()
        return user_groups

    def append_user_details(output_path, data):
        """Appends user details to an existing Excel file or creates one if it doesn't exist."""
        file_exists = os.path.exists(output_path)
        if file_exists:
            workbook = openpyxl.load_workbook(output_path)
            sheet = workbook.active
        else:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["First Name", "Last Name", "Email", "Phone", "Role", "User Group"])

        for row in data:
            sheet.append(row)
        workbook.save(output_path)

    def scrape_user_details(driver, user_group):
        """Scrapes user details from the current table, handles pagination."""
        data = []
        last_page_data = None
        while True:
            # Locate the user details table and scrape data
            rows = driver.find_elements(By.XPATH, "//div[@class='user-search']//tr")
            current_page_data = []
            for row in rows[1:]:
                cols = row.find_elements(By.TAG_NAME, "td")
                if cols:
                    first_name = cols[0].text
                    last_name = cols[1].text
                    email = cols[2].text
                    phone = cols[3].text
                    role = cols[4].text
                    current_page_data.append([first_name, last_name, email, phone, role, user_group])

            # Check if data has changed
            if current_page_data == last_page_data:
                time.sleep(DATA_CHECK_INTERVAL)
                if current_page_data == last_page_data:
                    break
            else:
                last_page_data = current_page_data
                data.extend(current_page_data)

            # Check if "Next" button is enabled
            try:
                next_button = driver.find_element(By.XPATH, "//i[@class='icon pos arrow-right']")
                if "disabled" in next_button.get_attribute("class"):
                    break
                next_button.click()
                time.sleep(2)
            except:
                break
        return data

    def open_all_user_groups(driver, user_group):
        """Opens all user groups listed in search results and scrapes details."""
        user_data = []
        results = driver.find_elements(By.XPATH, "//i[@class='icon pos view']")
        for i in range(len(results)):
            # Refresh the list of results
            results = driver.find_elements(By.XPATH, "//i[@class='icon pos view']")
            results[i].click()
            time.sleep(2)

            # Scrape user details, including pagination
            user_data.extend(scrape_user_details(driver, user_group))

            # Go back to the main page
            back_button = driver.find_element(By.XPATH, "//a[normalize-space()='Authorizer']")
            back_button.click()
            time.sleep(4)

        return user_data

    def main():
        while True:
            try:
                # Load user groups from Excel
                try:
                    user_groups = read_user_groups(INPUT_EXCEL_PATH)
                except FileNotFoundError:
                    print(f"Error: The file '{INPUT_EXCEL_PATH}' was not found.")
                    return

                # Initialize the Edge WebDriver
                options = Options()
                options.add_argument("start-maximized")
                service = Service("path/to/msedgedriver.exe")
                driver = webdriver.Edge(service=service, options=options)
                driver.implicitly_wait(IMPLICIT_WAIT)

                try:
                    driver.get(URL)

                    # Click on the User Groups button to start the script automatically
                    user_groups_button = WebDriverWait(driver, IMPLICIT_WAIT).until(
                        EC.presence_of_element_located((By.XPATH, "//a[normalize-space()='User Groups']"))
                    )
                    user_groups_button.click()

                    total_user_groups = len(user_groups)
                    for index, user_group in enumerate(user_groups):
                        print(f"{index + 1} out of {total_user_groups} user groups scraped")

                        # Clear the search box before entering a new user group name
                        search_box_found = False
                        while not search_box_found:
                            try:
                                search_box = driver.find_element(By.XPATH,
                                                                 "//input[@placeholder='Search User Group by Name or ID']")
                                search_box_found = True
                            except NoSuchElementException:
                                time.sleep(2)

                        search_box.click()
                        search_box.send_keys(Keys.CONTROL, "a")
                        search_box.send_keys(Keys.BACKSPACE)

                        # Enter the new user group name
                        search_box.send_keys(user_group)
                        search_box.send_keys(Keys.RETURN)
                        time.sleep(3)

                        # Open all user groups listed in search results and scrape details
                        user_data = open_all_user_groups(driver, user_group)
                        append_user_details(OUTPUT_EXCEL_PATH, user_data)

                finally:
                    driver.quit()
                    print("Scraping complete. Data saved to", OUTPUT_EXCEL_PATH)

                break

            except (NoSuchElementException, TimeoutException) as e:
                print(f"Error occurred: {e}. Restarting script...")

    if __name__ == "__main__":
        main()

    pass

def main():
    while True:
        print("Select a scraping option:")
        print("1) User group scraping (required for user detail scraping)")
        print("2) User details scraping (max 300 user groups)")
        print("3) User details scraping (more than 300 user groups, takes more time)")
        print("0) Exit")

        choice = input("Enter your choice: ").strip()
        if choice == "1":
            print("Running user group scraping...")
            script_option_1()
        elif choice == "2":
            print("Running user details scraping for max 300 user groups...")
            script_option_2()
        elif choice == "3":
            print("Running user details scraping for more than 300 user groups...")
            script_option_3()
        elif choice == "0":
            print("Exiting...")
            break
        else:
            print("Invalid choice. Please try again.")

if __name__ == "__main__":
    main()
